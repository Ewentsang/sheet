#!/usr/bin/env python3
"""
对比测试：展示美观效果与原始效果的差异
"""

import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_old_style_excel():
    """创建原始样式的Excel（用于对比）"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # 测试数据
    data = {
        "员工信息": [
            {"姓名": "张三", "部门": "技术部", "职位": "工程师", "薪资": 15000},
            {"姓名": "李四", "部门": "设计部", "职位": "设计师", "薪资": 12000},
            {"姓名": "王五", "部门": "市场部", "职位": "专员", "薪资": 8000}
        ]
    }
    
    for sheet_name, sheet_data in data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        if isinstance(sheet_data, list) and len(sheet_data) > 0:
            if isinstance(sheet_data[0], dict):
                headers = list(sheet_data[0].keys())
                # 原始样式：简单的灰色表头
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 写入数据（无样式）
                for row, row_data in enumerate(sheet_data, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=row_data.get(header, ""))
    
    # 简单的列宽调整
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def test_comparison():
    """对比测试"""
    print("=== 美观效果对比测试 ===")
    
    # 测试数据
    test_data = {
        "员工信息对比": [
            {"姓名": "张三", "部门": "技术开发部", "职位": "高级软件工程师", "薪资": 15000, "备注": "技术能力强，负责核心模块开发"},
            {"姓名": "李四", "部门": "产品设计部", "职位": "UI/UX设计师", "薪资": 12000, "备注": "擅长用户界面设计，有丰富的移动端经验"},
            {"姓名": "王五", "部门": "市场营销部", "职位": "市场推广专员", "薪资": 8000, "备注": "负责线上推广活动策划，熟悉社交媒体营销"},
            {"姓名": "赵六", "部门": "人力资源部", "职位": "HR专员", "薪资": 9000, "备注": "负责招聘和员工关系管理，沟通能力强"},
            {"姓名": "孙七", "部门": "财务部", "职位": "会计", "薪资": 10000, "备注": "负责公司财务核算和报表编制，工作细致认真"}
        ]
    }
    
    try:
        # 测试新版本（美观样式）
        print("📤 生成美观样式Excel...")
        response = requests.post(
            "http://localhost:5000/make-xlsx-bytes",
            json=test_data,
            headers={"Content-Type": "application/json"},
            timeout=30
        )
        
        if response.status_code == 200:
            with open("beautiful_style.xlsx", "wb") as f:
                f.write(response.content)
            print("✅ 美观样式Excel生成成功: beautiful_style.xlsx")
        else:
            print(f"❌ 美观样式生成失败: {response.status_code}")
            return
        
        # 生成原始样式对比
        print("📤 生成原始样式Excel...")
        old_wb = create_old_style_excel()
        old_wb.save("old_style.xlsx")
        print("✅ 原始样式Excel生成成功: old_style.xlsx")
        
        print("\n🎨 美观效果对比:")
        print("=" * 50)
        print("📊 原始样式特点:")
        print("   • 表头：浅灰色背景，简单加粗")
        print("   • 数据：无边框，无对齐，无条纹")
        print("   • 列宽：简单字符长度计算")
        print("   • 行高：默认行高")
        
        print("\n✨ 美观样式特点:")
        print("   • 表头：深蓝底(#1F4E79) + 白字 + 加粗，行高22")
        print("   • 斑马条纹：偶数行浅灰背景(#F2F2F2)")
        print("   • 边框：细边框包围所有单元格")
        print("   • 对齐：垂直居中，文本自动换行")
        print("   • 列宽：中英文智能计算，8-40字符范围")
        
        print("\n📁 生成的文件:")
        print("   • beautiful_style.xlsx - 美观样式版本")
        print("   • old_style.xlsx - 原始样式版本")
        print("\n💡 请打开两个文件对比查看效果差异！")
        
        return True
        
    except Exception as e:
        print(f"❌ 对比测试失败: {e}")
        return False

def main():
    """主函数"""
    print("🔄 Excel样式对比测试")
    print("=" * 50)
    
    # 检查服务状态
    try:
        response = requests.get("http://localhost:5000/health")
        if response.status_code == 200:
            print("✅ 服务运行正常")
        else:
            print("❌ 服务异常，请先启动服务")
            return
    except Exception as e:
        print(f"❌ 无法连接到服务: {e}")
        print("💡 请先运行: python app.py")
        return
    
    # 运行对比测试
    test_comparison()

if __name__ == "__main__":
    main()

