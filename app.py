from flask import Flask, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import io
import os
from datetime import datetime, timedelta
import boto3
from botocore.exceptions import ClientError
import logging
from pathlib import Path
from dotenv import load_dotenv

# 确保加载的是和 app.py 同一目录下的 .env
load_dotenv(dotenv_path=Path(__file__).with_name('.env'))



# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

@app.route("/ping", methods=["GET"])
def ping():
    return jsonify({"pong": True})
    
@app.route("/debug/env", methods=["GET"])
def debug_env():
    return jsonify({
        "S3_ENDPOINT_URL": os.getenv("S3_ENDPOINT_URL"),
        "S3_BUCKET": os.getenv("S3_BUCKET"),
        "S3_ACCESS_KEY": "****" if os.getenv("S3_ACCESS_KEY") else None,
        "S3_SECRET_KEY": "****" if os.getenv("S3_SECRET_KEY") else None,
        "S3_REGION": os.getenv("S3_REGION"),
        "URL_EXPIRY_HOURS": os.getenv("URL_EXPIRY_HOURS")
    })

# 配置
S3_BUCKET = os.getenv('S3_BUCKET', 'excel-files')
S3_REGION = os.getenv('S3_REGION', 'us-east-1')
S3_ACCESS_KEY = os.getenv('S3_ACCESS_KEY')
S3_SECRET_KEY = os.getenv('S3_SECRET_KEY')
S3_ENDPOINT_URL = os.getenv('S3_ENDPOINT_URL')  # 用于MinIO
URL_EXPIRY_HOURS = int(os.getenv('URL_EXPIRY_HOURS', '24'))

def create_excel_from_json(data):
    """
    从JSON数据创建Excel工作簿
    支持多sheet，每个sheet对应JSON中的一个键
    """
    wb = Workbook()
    
    # 移除默认的Sheet
    wb.remove(wb.active)
    
    if isinstance(data, dict):
        # 如果是字典，每个键创建一个sheet
        for sheet_name, sheet_data in data.items():
            create_sheet(wb, sheet_name, sheet_data)
    elif isinstance(data, list):
        # 如果是列表，创建一个默认sheet
        create_sheet(wb, "Sheet1", data)
    else:
        # 如果是单个值，创建一个默认sheet
        create_sheet(wb, "Sheet1", [{"value": data}])
    
    return wb

def detect_header_rows(data):
    """
    智能检测表头行数
    通过分析数据模式来区分表头和数据行
    """
    if not data or len(data) < 2:
        return min(1, len(data))
    
    # 方法1: 检测表头关键词（最可靠的方法）
    # 如果某行包含明显的表头关键词，这一行是最后一行表头
    for i, row in enumerate(data):
        row_str = ' '.join(str(cell) for cell in row if cell)
        if any(keyword in row_str.lower() for keyword in 
               ['agent_name', 'status_num', 'percentage']):
            return i + 1  # 这一行是表头，下一行开始是数据
    
    # 方法2: 检测第一列的数据类型模式
    # 寻找第一个看起来像具体数据而不是表头的行
    for i, row in enumerate(data):
        if row and len(row) > 0:
            first_cell = row[0]
            if first_cell and isinstance(first_cell, str):
                # 如果第一列开始出现人名（通常包含空格，且长度适中）
                if (' ' in first_cell and 5 <= len(first_cell) <= 20 and 
                    not any(keyword in first_cell.lower() for keyword in 
                           ['broadcast', 'open', 'solved', 'expired', 'status', 'percentage'])):
                    return max(1, i)
    
    # 方法3: 检测数值密度突然变化
    numeric_density = []
    for row in data:
        if not row:
            numeric_density.append(0)
            continue
        
        numeric_count = 0
        total_count = 0
        for cell in row:
            if cell and str(cell).strip():
                total_count += 1
                cell_str = str(cell).replace('%', '').replace(',', '')
                try:
                    float(cell_str)
                    numeric_count += 1
                except:
                    pass
        
        density = numeric_count / total_count if total_count > 0 else 0
        numeric_density.append(density)
    
    # 寻找数值密度突然增加的点
    for i in range(1, len(numeric_density)):
        if numeric_density[i] > numeric_density[i-1] + 0.2:  # 密度增加
            return i
    
    # 方法4: 检测空行分隔
    for i, row in enumerate(data):
        if not any(cell for cell in row if cell and str(cell).strip()):
            return max(1, i)
    
    # 默认策略：前3行作为表头
    return min(3, len(data) // 2)

def create_sheet(wb, sheet_name, data):
    """创建单个sheet"""
    # 清理sheet名称（Excel sheet名称限制）
    safe_name = sheet_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
    
    ws = wb.create_sheet(title=safe_name)
    
    # 定义样式
    # 表头样式：深蓝底 + 白字 + 加粗
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式：细边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 数据行样式：垂直居中，文本自动换行
    data_alignment = Alignment(vertical="center", wrap_text=True)
    
    # 斑马条纹样式
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    if isinstance(data, list) and len(data) > 0:
        # 如果是列表，检查第一行是否为字典（单层表头）还是列表（多层表头）
        if isinstance(data[0], dict):
            # 单层表头：字典列表格式
            headers = list(data[0].keys())
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 设置表头行高
            ws.row_dimensions[1].height = 22
            
            # 写入数据
            for row, row_data in enumerate(data, 2):
                # 斑马条纹：偶数行添加浅灰背景
                row_fill = zebra_fill if row % 2 == 0 else None
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=row_data.get(header, ""))
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
        elif isinstance(data[0], list):
            # 多层表头：二维数组格式
            # 智能识别表头行数：通过分析数据模式来区分表头和数据
            header_rows = detect_header_rows(data)
            
            # 写入多层表头（保持深蓝色表头样式）
            for header_row_idx in range(header_rows):
                row_data = data[header_row_idx]
                for col, cell_value in enumerate(row_data, 1):
                    # 确保单元格值是字符串或数字，不是列表
                    if isinstance(cell_value, list):
                        cell_value = str(cell_value)
                    elif cell_value is None:
                        cell_value = ""
                    cell = ws.cell(row=header_row_idx + 1, column=col, value=cell_value)
                    # 应用表头样式：深蓝底 + 白字 + 加粗
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # 设置表头行高（保持一致）
                ws.row_dimensions[header_row_idx + 1].height = 22
            
            # 写入数据行（保持斑马条纹样式）
            data_start_row = header_rows + 1
            for row_idx, row_data in enumerate(data[header_rows:], data_start_row):
                # 斑马条纹：偶数行添加浅灰背景（与单层表头完全相同的逻辑）
                row_fill = zebra_fill if row_idx % 2 == 0 else None
                
                for col, cell_value in enumerate(row_data, 1):
                    # 确保单元格值是字符串或数字，不是列表
                    if isinstance(cell_value, list):
                        cell_value = str(cell_value)
                    elif cell_value is None:
                        cell_value = ""
                    cell = ws.cell(row=row_idx, column=col, value=cell_value)
                    # 应用数据样式：垂直居中，文本自动换行，细边框
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
        else:
            # 如果是简单列表，直接写入
            for row, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=1, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border
                # 斑马条纹
                if row % 2 == 0:
                    cell.fill = zebra_fill
    elif isinstance(data, dict):
        # 如果是字典，创建键值对格式
        headers = ["Key", "Value"]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置表头行高
        ws.row_dimensions[1].height = 22
        
        # 写入数据
        for row, (key, value) in enumerate(data.items(), 2):
            # 斑马条纹
            row_fill = zebra_fill if row % 2 == 0 else None
            
            # Key列
            key_cell = ws.cell(row=row, column=1, value=key)
            key_cell.alignment = data_alignment
            key_cell.border = thin_border
            if row_fill:
                key_cell.fill = row_fill
            
            # Value列
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.alignment = data_alignment
            value_cell.border = thin_border
            if row_fill:
                value_cell.fill = row_fill
    else:
        # 单个值
        # 表头
        header_cell = ws.cell(row=1, column=1, value="Value")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        ws.row_dimensions[1].height = 22
        
        # 数据
        data_cell = ws.cell(row=2, column=1, value=data)
        data_cell.alignment = data_alignment
        data_cell.border = thin_border
    
    # 智能列宽调整：根据中英文宽度估算
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                # 计算字符宽度：中文字符算2个宽度，英文算1个宽度
                width = 0
                for char in cell_value:
                    if ord(char) > 127:  # 中文字符
                        width += 2
                    else:  # 英文字符
                        width += 1
                
                if width > max_length:
                    max_length = width
            except:
                pass
        
        # 列宽控制在8-40字符范围，并添加适当边距
        adjusted_width = max(8, min(max_length + 3, 40))
        ws.column_dimensions[column_letter].width = adjusted_width

def copy_file_with_new_name(source_file_path, new_filename):
    """
    复制文件并重命名
    
    Args:
        source_file_path (str): 源文件路径
        new_filename (str): 新的文件名
    
    Returns:
        str: 新文件的完整路径
    
    Raises:
        FileNotFoundError: 源文件不存在
        Exception: 复制过程中发生错误
    """
    try:
        # 检查源文件是否存在
        if not os.path.exists(source_file_path):
            raise FileNotFoundError(f"源文件不存在: {source_file_path}")
        
        # 获取源文件所在目录
        source_dir = os.path.dirname(source_file_path)
        
        # 构建新文件的完整路径
        new_file_path = os.path.join(source_dir, new_filename)
        
        # 如果新文件已存在，先删除
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
            logger.info(f"已删除已存在的文件: {new_file_path}")
        
        # 复制文件
        import shutil
        shutil.copy2(source_file_path, new_file_path)
        
        logger.info(f"文件复制成功: {source_file_path} -> {new_file_path}")
        return new_file_path
        
    except FileNotFoundError as e:
        logger.error(f"文件不存在错误: {e}")
        raise
    except Exception as e:
        logger.error(f"文件复制错误: {e}")
        raise Exception(f"文件复制失败: {str(e)}")

def upload_to_s3(excel_bytes, filename):
    """上传Excel文件到S3/MinIO"""
    try:
        if S3_ENDPOINT_URL:
            # MinIO配置
            s3_client = boto3.client(
                's3',
                endpoint_url=S3_ENDPOINT_URL,
                aws_access_key_id=S3_ACCESS_KEY,
                aws_secret_access_key=S3_SECRET_KEY,
                region_name=S3_REGION
            )
        else:
            # AWS S3配置
            s3_client = boto3.client(
                's3',
                aws_access_key_id=S3_ACCESS_KEY,
                aws_secret_access_key=S3_SECRET_KEY,
                region_name=S3_REGION
            )
        
        # 上传文件
        s3_client.put_object(
            Bucket=S3_BUCKET,
            Key=filename,
            Body=excel_bytes,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 生成预签名URL
        presigned_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': S3_BUCKET, 'Key': filename},
            ExpiresIn=URL_EXPIRY_HOURS * 3600
        )
        
        return presigned_url
        
    except ClientError as e:
        logger.error(f"S3上传错误: {e}")
        raise Exception(f"S3上传失败: {str(e)}")
    except Exception as e:
        logger.error(f"上传错误: {e}")
        raise Exception(f"上传失败: {str(e)}")

@app.route('/health', methods=['GET'])
def health_check():
    """健康检查端点"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

@app.route('/make-xlsx-bytes', methods=['POST'])
def make_xlsx_bytes():
    """直接返回Excel二进制文件流"""
    try:
        # 获取JSON数据
        if request.is_json:
            data = request.get_json()
        else:
            # 如果不是JSON，尝试解析表单数据
            data_str = request.form.get('data') or request.data.decode('utf-8')
            data = json.loads(data_str)
        
        # 创建Excel
        wb = create_excel_from_json(data)
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except json.JSONDecodeError as e:
        return jsonify({"error": f"JSON解析错误: {str(e)}"}), 400
    except Exception as e:
        logger.error(f"生成Excel错误: {e}")
        return jsonify({"error": f"生成Excel失败: {str(e)}"}), 500

@app.route('/make-xlsx-url', methods=['POST'])
def make_xlsx_url():
    """生成Excel并上传到S3，返回下载URL"""
    try:
        # 获取JSON数据
        if request.is_json:
            data = request.get_json()
        else:
            # 如果不是JSON，尝试解析表单数据
            data_str = request.form.get('data') or request.data.decode('utf-8')
            data = json.loads(data_str)
        
        # 检查S3配置
        if not S3_ACCESS_KEY or not S3_SECRET_KEY:
            return jsonify({"error": "S3配置缺失"}), 500
        
        # 创建Excel
        wb = create_excel_from_json(data)
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.getvalue()
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_{timestamp}.xlsx"
        
        # 上传到S3
        download_url = upload_to_s3(excel_bytes, filename)
        
        return jsonify({
            "success": True,
            "download_url": download_url,
            "filename": filename,
            "expires_in_hours": URL_EXPIRY_HOURS,
            "expires_at": (datetime.now() + timedelta(hours=URL_EXPIRY_HOURS)).isoformat()
        })
        
    except json.JSONDecodeError as e:
        return jsonify({"error": f"JSON解析错误: {str(e)}"}), 400
    except Exception as e:
        logger.error(f"生成Excel URL错误: {e}")
        return jsonify({"error": f"生成Excel URL失败: {str(e)}"}), 500

@app.route('/copy-file', methods=['POST'])
def copy_file():
    """上传文件并重命名，直接返回重命名后的文件"""
    try:
        logger.info("收到文件上传请求")
        
        # 检查是否有文件上传
        if 'file' not in request.files:
            logger.error("没有上传文件")
            return jsonify({"error": "没有上传文件"}), 400
        
        uploaded_file = request.files['file']
        logger.info(f"上传文件: {uploaded_file.filename}")
        
        # 检查文件名是否为空
        if uploaded_file.filename == '':
            logger.error("没有选择文件")
            return jsonify({"error": "没有选择文件"}), 400
        
        # 获取新文件名
        new_filename = request.form.get('new_filename')
        logger.info(f"新文件名: {new_filename}")
        
        if not new_filename:
            logger.error("缺少参数: new_filename")
            return jsonify({"error": "缺少参数: new_filename"}), 400
        
        # 确保新文件名有扩展名
        if not new_filename.endswith(('.xlsx', '.xls', '.csv', '.txt', '.pdf', '.doc', '.docx')):
            # 如果没有扩展名，使用原文件的扩展名
            original_ext = os.path.splitext(uploaded_file.filename)[1]
            new_filename = new_filename + original_ext
            logger.info(f"添加扩展名后的文件名: {new_filename}")
        
        # 保存上传的文件到临时位置
        temp_dir = "temp_uploads"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        
        temp_file_path = os.path.join(temp_dir, uploaded_file.filename)
        uploaded_file.save(temp_file_path)
        logger.info(f"文件保存到临时位置: {temp_file_path}")
        
        new_file_path = None
        try:
            # 调用复制函数
            new_file_path = copy_file_with_new_name(temp_file_path, new_filename)
            logger.info(f"文件复制成功: {new_file_path}")
            
            # 直接返回文件，就像 make_xlsx_bytes 一样
            # 使用 Response 对象和 after_request 来延迟清理
            response = send_file(
                new_file_path,
                as_attachment=True,
                download_name=new_filename
            )
            
            # 在响应完成后清理文件
            def cleanup_after_response():
                try:
                    # 清理临时文件
                    if os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                        logger.info(f"已清理临时文件: {temp_file_path}")
                    
                    # 清理复制后的文件
                    if new_file_path and os.path.exists(new_file_path):
                        os.remove(new_file_path)
                        logger.info(f"已清理复制文件: {new_file_path}")
                except Exception as e:
                    logger.warning(f"清理文件时出现警告: {e}")
            
            # 将清理函数附加到响应对象
            response.call_on_close(cleanup_after_response)
            return response
            
        except Exception as e:
            # 如果出现异常，立即清理文件
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                    logger.info(f"异常时清理临时文件: {temp_file_path}")
                
                if new_file_path and os.path.exists(new_file_path):
                    os.remove(new_file_path)
                    logger.info(f"异常时清理复制文件: {new_file_path}")
            except Exception as cleanup_error:
                logger.warning(f"异常清理文件时出现警告: {cleanup_error}")
            raise
        
    except FileNotFoundError as e:
        logger.error(f"文件不存在错误: {e}")
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logger.error(f"文件上传错误: {e}")
        return jsonify({"error": f"文件上传失败: {str(e)}"}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "接口不存在"}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "服务器内部错误"}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5014))
    app.run(host='0.0.0.0', port=port, debug=False)


